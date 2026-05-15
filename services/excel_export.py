"""
services/excel_export.py — генерация Excel-отчётов через openpyxl.

Создаёт .xlsx файл с тремя листами:
  1. «Транзакции»  — все операции за период
  2. «По категориям» — сводная таблица с формулами
  3. «Итоги»       — доходы, расходы, баланс по месяцам
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# Цвета
CLR_HEADER_BG  = "1F4E79"   # тёмно-синий — шапка
CLR_HEADER_FG  = "FFFFFF"   # белый текст
CLR_INCOME_BG  = "E2EFDA"   # светло-зелёный — доходы
CLR_EXPENSE_BG = "FCE4D6"   # светло-красный — расходы
CLR_TOTAL_BG   = "D9E1F2"   # голубой — итоги
CLR_ALT_ROW    = "F2F2F2"   # серый — чётные строки
CLR_ACCENT     = "2E75B6"   # синий — акцент

#  Стили
FONT_MAIN   = "Arial"
FONT_MONO   = "Courier New"

def _hdr_font(size=11):
    return Font(name=FONT_MAIN, bold=True, color=CLR_HEADER_FG, size=size)

def _bold(size=10):
    return Font(name=FONT_MAIN, bold=True, size=size)

def _normal(size=10):
    return Font(name=FONT_MAIN, size=size)

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

def _right():
    return Alignment(horizontal="right", vertical="center")


def _set_cell(ws, row, col, value, font=None, fill=None,
              alignment=None, number_format=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if font:       cell.font = font
    if fill:       cell.fill = fill
    if alignment:  cell.alignment = alignment
    if number_format: cell.number_format = number_format
    if border:     cell.border = _border()
    return cell



# ГЛАВНАЯ ФУНКЦИЯ


def generate_excel_report(
    transactions: list[dict],
    year: int,
    month: int | None = None,
    currency: str = "с",
) -> bytes:
    """
    Генерирует Excel-отчёт и возвращает bytes.

    transactions: список словарей из get_all_transactions_for_export()
    year: год отчёта
    month: месяц (None = весь год)
    currency: символ валюты
    """
    wb = Workbook()
    wb.remove(wb.active)  # удаляем дефолтный лист

    # Фильтруем транзакции по периоду
    period_str = f"{year}-{month:02d}" if month else str(year)
    filtered = [
        t for t in transactions
        if t["txn_date"].startswith(period_str)
    ]

    _sheet_transactions(wb, filtered, currency, year, month)
    _sheet_by_category(wb, filtered, currency)
    _sheet_summary(wb, transactions, currency, year)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()



# ЛИСТ 1: ТРАНЗАКЦИИ


def _sheet_transactions(wb, transactions, currency, year, month):
    ws = wb.create_sheet("Транзакции")
    ws.sheet_view.showGridLines = True

    # Заголовок
    period_label = f"{_month_name(month)} {year}" if month else f"{year} год"
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"Финансовый отчёт — {period_label}"
    title.font = Font(name=FONT_MAIN, bold=True, size=14, color=CLR_ACCENT)
    title.alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    sub.font = Font(name=FONT_MAIN, size=9, color="888888")
    sub.alignment = _center()
    ws.row_dimensions[2].height = 16

    # Шапка таблицы
    headers = ["№", "Дата", "Тип", "Категория", "Описание", f"Сумма ({currency})", "Источник"]
    col_widths = [5, 12, 10, 18, 30, 14, 10]

    for col, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        _set_cell(ws, 4, col, hdr,
                  font=_hdr_font(10),
                  fill=_fill(CLR_HEADER_BG),
                  alignment=_center())
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 22

    # Данные
    sorted_txns = sorted(transactions, key=lambda x: x["txn_date"], reverse=True)
    income_total = 0.0
    expense_total = 0.0

    for i, t in enumerate(sorted_txns, 1):
        row = i + 4
        is_income = bool(t["is_income"])
        bg = CLR_INCOME_BG if is_income else (CLR_EXPENSE_BG if i % 2 == 0 else "FFFFFF")

        row_fill = _fill(bg)
        row_font = _normal(10)

        src_label = {"voice": "🎙 Голос", "text": "✏️ Текст"}.get(t.get("source", "text"), "✏️ Текст")

        _set_cell(ws, row, 1, i,                          font=row_font, fill=row_fill, alignment=_center())
        _set_cell(ws, row, 2, t["txn_date"],              font=row_font, fill=row_fill, alignment=_center())
        _set_cell(ws, row, 3, "Доход" if is_income else "Расход",
                                                           font=row_font, fill=row_fill, alignment=_center())
        _set_cell(ws, row, 4, t.get("category") or "—",  font=row_font, fill=row_fill, alignment=_left())
        _set_cell(ws, row, 5, t.get("note") or "—",      font=row_font, fill=row_fill, alignment=_left())
        amt_cell = _set_cell(ws, row, 6, t["amount"],     font=row_font, fill=row_fill, alignment=_right(),
                             number_format='#,##0.00')
        _set_cell(ws, row, 7, src_label,                  font=row_font, fill=row_fill, alignment=_center())

        if is_income:
            income_total += t["amount"]
        else:
            expense_total += t["amount"]

    # Строка итогов
    last_data_row = len(sorted_txns) + 4
    total_row = last_data_row + 2

    ws.merge_cells(f"A{total_row}:E{total_row}")
    _set_cell(ws, total_row, 1, "ИТОГО:", font=_bold(11),
              fill=_fill(CLR_TOTAL_BG), alignment=_right(), border=False)

    income_row = total_row + 1
    expense_row = total_row + 2
    balance_row = total_row + 3

    for r, label, val, clr in [
        (income_row,  "💚 Доходы:",  income_total,              CLR_INCOME_BG),
        (expense_row, "💸 Расходы:", expense_total,             CLR_EXPENSE_BG),
        (balance_row, "📊 Баланс:",  income_total - expense_total, CLR_TOTAL_BG),
    ]:
        ws.merge_cells(f"A{r}:E{r}")
        _set_cell(ws, r, 1, label, font=_bold(10),
                  fill=_fill(clr), alignment=_right(), border=False)
        _set_cell(ws, r, 6, val, font=_bold(10),
                  fill=_fill(clr), alignment=_right(),
                  number_format='#,##0.00', border=True)

    ws.freeze_panes = "A5"



# ЛИСТ 2: ПО КАТЕГОРИЯМ


def _sheet_by_category(wb, transactions, currency):
    ws = wb.create_sheet("По категориям")

    # Агрегация
    cat_data: dict[str, dict] = {}
    total_exp = 0.0
    for t in transactions:
        if t["is_income"]:
            continue
        cat = t.get("category") or "Без категории"
        if cat not in cat_data:
            cat_data[cat] = {"total": 0.0, "count": 0}
        cat_data[cat]["total"] += t["amount"]
        cat_data[cat]["count"] += 1
        total_exp += t["amount"]

    sorted_cats = sorted(cat_data.items(), key=lambda x: x[1]["total"], reverse=True)

    # Заголовок
    ws.merge_cells("A1:E1")
    ws["A1"].value = "Расходы по категориям"
    ws["A1"].font = Font(name=FONT_MAIN, bold=True, size=13, color=CLR_ACCENT)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["Категория", "Операций", f"Сумма ({currency})", "Доля %", "Визуализация"]
    widths   = [25, 12, 16, 12, 30]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _set_cell(ws, 3, col, h, font=_hdr_font(10),
                  fill=_fill(CLR_HEADER_BG), alignment=_center())
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    for i, (cat, data) in enumerate(sorted_cats, 1):
        row = i + 3
        bg = _fill(CLR_ALT_ROW) if i % 2 == 0 else _fill("FFFFFF")
        pct = data["total"] / total_exp * 100 if total_exp else 0
        bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))

        _set_cell(ws, row, 1, cat,           font=_normal(), fill=bg, alignment=_left())
        _set_cell(ws, row, 2, data["count"], font=_normal(), fill=bg, alignment=_center())
        _set_cell(ws, row, 3, data["total"], font=_normal(), fill=bg,
                  alignment=_right(), number_format='#,##0.00')
        _set_cell(ws, row, 4, round(pct, 1), font=_normal(), fill=bg,
                  alignment=_center(), number_format='0.0"%"')
        _set_cell(ws, row, 5, f"{bar} {pct:.1f}%",
                  font=Font(name=FONT_MONO, size=9), fill=bg, alignment=_left())

    # Итого
    last = len(sorted_cats) + 4
    ws.merge_cells(f"A{last}:B{last}")
    _set_cell(ws, last, 1, "ИТОГО РАСХОДОВ:", font=_bold(10),
              fill=_fill(CLR_TOTAL_BG), alignment=_right(), border=True)
    _set_cell(ws, last, 3, total_exp, font=_bold(10),
              fill=_fill(CLR_TOTAL_BG), alignment=_right(),
              number_format='#,##0.00')
    ws.freeze_panes = "A4"



# ЛИСТ 3: ИТОГИ ПО МЕСЯЦАМ


def _sheet_summary(wb, transactions, currency, year):
    ws = wb.create_sheet("Итоги по месяцам")

    MONTHS = ["Январь","Февраль","Март","Апрель","Май","Июнь",
              "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]

    # Агрегация по месяцам
    monthly: dict[int, dict] = {m: {"income": 0.0, "expense": 0.0} for m in range(1, 13)}
    for t in transactions:
        if not t["txn_date"].startswith(str(year)):
            continue
        try:
            m = int(t["txn_date"][5:7])
            if t["is_income"]:
                monthly[m]["income"] += t["amount"]
            else:
                monthly[m]["expense"] += t["amount"]
        except (ValueError, IndexError):
            continue

    # Заголовок
    ws.merge_cells("A1:E1")
    ws["A1"].value = f"Итоги по месяцам — {year} год"
    ws["A1"].font = Font(name=FONT_MAIN, bold=True, size=13, color=CLR_ACCENT)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    headers = ["Месяц", f"Доходы ({currency})", f"Расходы ({currency})",
               f"Баланс ({currency})", "Статус"]
    widths   = [16, 16, 16, 16, 14]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _set_cell(ws, 3, col, h, font=_hdr_font(10),
                  fill=_fill(CLR_HEADER_BG), alignment=_center())
        ws.column_dimensions[get_column_letter(col)].width = w

    for m in range(1, 13):
        row = m + 3
        inc = monthly[m]["income"]
        exp = monthly[m]["expense"]
        bal = inc - exp
        status = "✅ Профицит" if bal > 0 else ("⚖️ Ноль" if bal == 0 else "❌ Дефицит")
        bg_bal = _fill(CLR_INCOME_BG) if bal > 0 else (_fill(CLR_EXPENSE_BG) if bal < 0 else _fill("FFFFFF"))
        row_bg = _fill(CLR_ALT_ROW) if m % 2 == 0 else _fill("FFFFFF")

        _set_cell(ws, row, 1, MONTHS[m-1], font=_bold(10),  fill=row_bg, alignment=_left())
        _set_cell(ws, row, 2, inc,         font=_normal(),  fill=_fill(CLR_INCOME_BG),
                  alignment=_right(), number_format='#,##0.00')
        _set_cell(ws, row, 3, exp,         font=_normal(),  fill=_fill(CLR_EXPENSE_BG),
                  alignment=_right(), number_format='#,##0.00')
        _set_cell(ws, row, 4, bal,         font=_bold(10),  fill=bg_bal,
                  alignment=_right(), number_format='#,##0.00')
        _set_cell(ws, row, 5, status,      font=_normal(),  fill=bg_bal, alignment=_center())

    # Годовой итог
    total_row = 16
    ws.merge_cells(f"A{total_row}:A{total_row}")
    total_inc = sum(monthly[m]["income"]  for m in range(1, 13))
    total_exp = sum(monthly[m]["expense"] for m in range(1, 13))
    total_bal = total_inc - total_exp

    _set_cell(ws, total_row, 1, "ГОД ИТОГО", font=_bold(11),
              fill=_fill(CLR_TOTAL_BG), alignment=_center())
    _set_cell(ws, total_row, 2, total_inc, font=_bold(11),
              fill=_fill(CLR_TOTAL_BG), alignment=_right(), number_format='#,##0.00')
    _set_cell(ws, total_row, 3, total_exp, font=_bold(11),
              fill=_fill(CLR_TOTAL_BG), alignment=_right(), number_format='#,##0.00')
    _set_cell(ws, total_row, 4, total_bal, font=_bold(11),
              fill=_fill(CLR_INCOME_BG if total_bal >= 0 else CLR_EXPENSE_BG),
              alignment=_right(), number_format='#,##0.00')
    _set_cell(ws, total_row, 5,
              "✅ Профицит" if total_bal >= 0 else "❌ Дефицит",
              font=_bold(10),
              fill=_fill(CLR_INCOME_BG if total_bal >= 0 else CLR_EXPENSE_BG),
              alignment=_center())
    ws.freeze_panes = "A4"


# Вспомогательное

def _month_name(month: int | None) -> str:
    names = ["","Январь","Февраль","Март","Апрель","Май","Июнь",
             "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"]
    return names[month] if month else ""